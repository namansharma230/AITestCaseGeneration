"""
excel_handler.py
Handles all Excel operations using openpyxl.

Features:
- Every run creates a FRESH sub-sheet — never appends to an existing one
- Sheet name = ticket ID + timestamp (e.g. "ALTV-551 2026-03-23 14:30")
- Styled headers, alternating row colours, frozen pane
"""

import logging
import re
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import EXCEL_FILE_PATH, EXCEL_HEADERS

logger = logging.getLogger(__name__)

# ── Styling ───────────────────────────────────────────────────────────────────
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL  = PatternFill(fill_type="solid", fgColor="2E4057")
_CELL_ALIGN   = Alignment(wrap_text=True, vertical="top")

_COLUMN_WIDTHS: dict[str, int] = {
    "Test Case ID":    14,
    "Title":           35,
    "Description":     40,
    "Preconditions":   30,
    "Steps":           50,
    "Expected Result": 40,
    "Postconditions":  30,
    "Priority":        12,
    "Platform":        18,
}

_INVALID_CHARS  = re.compile(r'[\\/?*\[\]:]')
_MAX_SHEET_NAME = 31


def sanitise_sheet_name(name: str) -> str:
    """Strip invalid characters and truncate to Excel's 31-char sheet name limit."""
    safe = _INVALID_CHARS.sub("", name).strip()
    return safe[:_MAX_SHEET_NAME] if safe else "Sheet"


def _unique_sheet_name(wb: Workbook, base_name: str) -> str:
    """
    Build a sheet name that does not already exist in *wb*.

    Strategy:
      1. Try  "<base_name> HH:MM"          e.g. "ALTV-551 14:30"
      2. If taken, try "<base_name> HH:MM:SS"  e.g. "ALTV-551 14:30:05"
      3. If still taken, append incrementing counter until free.

    This guarantees every run gets its own fresh sheet even if two runs
    happen within the same minute.

    Args:
        wb:        Open openpyxl Workbook.
        base_name: Sanitised ticket ID or requirement name.

    Returns:
        A sheet name string guaranteed not to exist in *wb*.
    """
    now = datetime.now()

    # Attempt 1 — base + HH:MM  (fits nicely in 31 chars for short ticket IDs)
    candidate = sanitise_sheet_name(f"{base_name} {now.strftime('%H:%M')}")
    if candidate not in wb.sheetnames:
        return candidate

    # Attempt 2 — base + HH:MM:SS
    candidate = sanitise_sheet_name(f"{base_name} {now.strftime('%H:%M:%S')}")
    if candidate not in wb.sheetnames:
        return candidate

    # Attempt 3 — append incrementing counter
    counter = 2
    while True:
        candidate = sanitise_sheet_name(f"{base_name} ({counter})")
        if candidate not in wb.sheetnames:
            return candidate
        counter += 1


def _write_headers(ws) -> None:
    """Write the styled header row and freeze pane on row 1."""
    for col_idx, header in enumerate(EXCEL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _CELL_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = (
            _COLUMN_WIDTHS.get(header, 20)
        )
    ws.freeze_panes = "A2"
    logger.debug("Headers written on sheet '%s'.", ws.title)


def _style_data_row(ws, row_num: int) -> None:
    """Apply wrap-text and alternating row background to a data row."""
    fill = PatternFill(
        fill_type="solid",
        fgColor="F2F4F7" if row_num % 2 == 0 else "FFFFFF"
    )
    for col_idx in range(1, len(EXCEL_HEADERS) + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.alignment = _CELL_ALIGN
        cell.fill      = fill


def append_test_cases(rows: list[list[str]], sheet_name: str = "Test Cases") -> None:
    """
    Write test-case rows to a BRAND NEW sub-sheet every time.

    The sheet name is built as "<ticket_id> HH:MM" so each run is clearly
    identifiable by time and no data is ever overwritten or appended to a
    previous run's sheet.

    Args:
        rows:       List of 7-element row lists
                    [ID, Title, Preconditions, Steps, Expected Result,
                     Postconditions, Priority].
        sheet_name: Base name (ticket ID / requirement name). A timestamp
                    suffix is added automatically to make it unique.

    Raises:
        ValueError: If rows is empty.
        IOError:    If the workbook cannot be saved (file open in Excel).
    """
    if not rows:
        raise ValueError("No rows provided to append_test_cases().")

    base_name = sanitise_sheet_name(sheet_name)
    path      = Path(EXCEL_FILE_PATH)

    # Load existing workbook or create a fresh one
    if path.exists():
        wb = openpyxl.load_workbook(path)
    else:
        path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # Always create a new sheet — never reuse an existing one
    final_name = _unique_sheet_name(wb, base_name)
    ws = wb.create_sheet(title=final_name)
    _write_headers(ws)
    logger.info("Created new sheet: '%s'", final_name)

    next_row = 2  # Row 1 is headers
    for row_data in rows:
        ws.append(row_data)
        _style_data_row(ws, next_row)
        next_row += 1

    try:
        wb.save(EXCEL_FILE_PATH)
        logger.info(
            "✓ Saved %d row(s) → '%s' (sheet: '%s')",
            len(rows), EXCEL_FILE_PATH, final_name,
        )
    except PermissionError:
        raise IOError(
            f"Cannot save '{EXCEL_FILE_PATH}'. "
            "Close the file in Excel and try again."
        )


def append_summary_to_excel(base_name: str, summary: dict, dependencies: list) -> None:
    """
    Write summary and dependencies to a new sheet in summary_requirements.xlsx.
    Creates a combined sheet with the Summary info at the top and the Dependencies table below.
    """
    summary_file_path = Path(EXCEL_FILE_PATH).parent / "summary_requirements.xlsx"
    clean_base = sanitise_sheet_name(base_name)

    if summary_file_path.exists():
        wb = openpyxl.load_workbook(summary_file_path)
    else:
        summary_file_path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    final_name = _unique_sheet_name(wb, clean_base)
    ws = wb.create_sheet(title=final_name)

    # ── Write Summary Section ──
    # Header
    ws.append(["--- REQUIREMENT SUMMARY ---"])
    cell = ws.cell(row=ws.max_row, column=1)
    cell.font = Font(bold=True, size=12, color="2E4057")

    # Data
    ws.append(["Overview", summary.get("overview", "")])
    ws.append(["Scope", summary.get("scope", "")])
    ws.append(["Complexity", summary.get("complexity", "")])

    features = summary.get("key_features", [])
    features_text = "\n".join(f"• {f}" for f in features)
    ws.append(["Key Features", features_text])
    
    # Style summary section
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).alignment = _CELL_ALIGN

    ws.append([])  # Blank row spacer

    # ── Write Dependencies Section ──
    ws.append(["--- TESTING DEPENDENCIES ---"])
    cell = ws.cell(row=ws.max_row, column=1)
    cell.font = Font(bold=True, size=12, color="2E4057")

    deps_headers = ["Category", "Item", "Description", "Owner", "Priority"]
    ws.append(deps_headers)
    
    # Style deps headers
    header_row = ws.max_row
    for col_idx in range(1, len(deps_headers) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CELL_ALIGN

    # Dependencies Data
    if not dependencies:
        ws.append(["No dependencies identified."])
    else:
        for i, dep in enumerate(dependencies):
            ws.append([
                dep.get("category", ""),
                dep.get("item", ""),
                dep.get("description", ""),
                dep.get("owner", ""),
                dep.get("priority", "")
            ])
            # Alternating row colors
            row_idx = ws.max_row
            fill = PatternFill(fill_type="solid", fgColor="F2F4F7" if i % 2 == 0 else "FFFFFF")
            for col_idx in range(1, len(deps_headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = _CELL_ALIGN
                cell.fill = fill

    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 15

    try:
        wb.save(summary_file_path)
        logger.info("✓ Saved summary & %d dependencies → 'summary_requirements.xlsx' (sheet: '%s')", len(dependencies), final_name)
    except PermissionError:
        raise IOError(f"Cannot save '{summary_file_path}'. Close the file in Excel and try again.")